# '21.09.11 토요일에 최초로 완성하고 1,000개 종목의 데이터를 가져와 저장해 봤음
# '21.12.03 토요일에 검토 유니버스를 1,500개로 늘렸음 (데이터 시트에 종목 1,500개만 넣으면 됨)
# '22.01.01 토요일에 OneDrive 폴더구조 변경하고 파일이름 변경해서, 관련내용 변경하였음.

#   1. 시가총액 순위별로 회사 Code가 입력되어 있는 파일을 참조할 준비를 한다
#   2. 그 파일에 불러온 값을 저장할 준비도 한다
from openpyxl import load_workbook
path = "C:/Users/msmk1/OneDrive/1 Routine/1.1 Automation_data/Data1500.xlsx"
#      "c:/Users/msmk1/OneDrive/Routine/Auto/market_cap_rank.xlsx"  - 220101 Update 전의 Code   
wb = load_workbook(path)
ws = wb["Rank"]
db = wb["Data"]


#   3. 필요한 라이브러리들 불러오고
import requests
from bs4 import BeautifulSoup


# 슬랙 메시지 전달
# def post_message(token, channel, text):
#    response = requests.post("https://slack.com/api/chat.postMessage",
#        headers={"Authorization": "Bearer "+token},
#        data={"channel": channel,"text": text}
#    )
#    print(response)
# myToken = "xoxb-2483926020578-2469328976551-Qcu5xUqVzxqSsPeQ0Mv5RxHz"


#   4. 회사 Code와 이름을 불러온다
for x in range(1,ws.max_row+1):
    company_Code = ws.cell(row=x, column=3).value
    company_Name = ws.cell(row=x, column=2).value

#   5. FnGuide 해당 페이지 접속해서 데이터 가져온다

#   url 주소가 약간 바뀌었음을 확인, 변경했음. '21.12.03
#   url = "https://comp.fnguide.com/SVO2/ASP/SVD_main.asp?pGB=1&gicode=A" + company_Code + "&cID=&MenuYn=Y&ReportGB=&NewMenuID=11&stkGb=&strResearchYN="
    url = "https://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode=A" + company_Code + "&&cID=&MenuYn=Y&ReportGB=&NewMenuID=101&stkGb=701"
    res = requests.get(url)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")


#   6. Financial Highlight 항목 명칭을 입력한다.
    data_rows = soup.find("div", attrs={"id":"highlight_D_Y"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    i = 0
    for row in data_rows:
        i = i + 1
        C_head = row.find("th").find("div").get_text()
        db.cell(row=i + 25 * (x-1), column=1, value= company_Name)
        db.cell(row=i + 25 * (x-1), column=5, value= C_head)


#   7. 우선주 입력한다.
    basic_info = soup.find("div", attrs={"id":"div1"}).find("table", attrs={"class":"us_table_ty1 table-hb thbg_g h_fix zigbg_no"}).find("tbody").find_all("tr")
    i = 0
    for row in basic_info:
        i = i + 1
        columns = row.find_all("td")
        j = 0
        for col in columns:
            j = j + 1
            data = col.get_text()
            if i == 7 and j == 1:
                F_data = data   
    db.cell(row=1 + 25 * (x-1), column=2, value= F_data)


#   8. 자사주 입력한다.
    basic_info = soup.find("div", attrs={"id":"svdMainGrid5"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no notres"}).find("tbody").find_all("tr")
    i = 0
    for row in basic_info:
        i = i + 1
        columns = row.find_all("td")
        j = 0
        for col in columns:
            j = j + 1
            data = col.get_text()
            if i == 5 and j == 2:
                F_data = data
    db.cell(row=1 + 25 * (x-1), column=3, value= F_data)


#   9. Financial Highlight, 연결/연간의 연도별 값을 입력한다.
    i = 0
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
# 여기 아래에서 에러가 발생한다면,
# 진행률 표시된 바로 다음 종목 데이터가 FnGuide 홈페이지에 없는 것임
# 그 종목 제외하고 List 만들어서 다시 돌리면 된다.
        db.cell(row=i + 25 * (x-1), column=6, value = data[0])
        db.cell(row=i + 25 * (x-1), column=7, value = data[1])
        db.cell(row=i + 25 * (x-1), column=8, value = data[2])
        db.cell(row=i + 25 * (x-1), column=9, value = data[3])
        db.cell(row=i + 25 * (x-1), column=10, value = data[4])
        db.cell(row=i + 25 * (x-1), column=11, value = data[5])
        db.cell(row=i + 25 * (x-1), column=12, value = data[6])
        db.cell(row=i + 25 * (x-1), column=13, value = data[7])


#   10. Financial Highlight, 연결/분기의 분기별 값을 입력한다.
    data_rows = soup.find("div", attrs={"id":"highlight_D_Q"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")

    i = 0
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        db.cell(row=i + 25 * (x-1), column=15, value = data[0])
        db.cell(row=i + 25 * (x-1), column=16, value = data[1])
        db.cell(row=i + 25 * (x-1), column=17, value = data[2])
        db.cell(row=i + 25 * (x-1), column=18, value = data[3])
        db.cell(row=i + 25 * (x-1), column=19, value = data[4])
        db.cell(row=i + 25 * (x-1), column=20, value = data[5])
        db.cell(row=i + 25 * (x-1), column=21, value = data[6])
        db.cell(row=i + 25 * (x-1), column=22, value = data[7])


#   11. 진행된 내용 표시하고, 200개 검토될 때마다 저장한다.
    print("진행률 : " + str(x) + "/" + str(ws.max_row+1) + ", " + company_Name)
    if x % 200 == 0:
            wb.save(path)


#   12. 다되면 저장한다.    
wb.save(path)

# post_message(myToken,"#stock_test","Done")