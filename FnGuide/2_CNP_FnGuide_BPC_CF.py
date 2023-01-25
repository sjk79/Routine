# '21.09.16 작성 시작해서 당일 끝내고, 1000종목 가져왔음.
# CNP_FnGuide_BPC에서 일부 항목만 가져오는 것으로 수정함.
# 매출, 영업이익, 당기순이익, 영업CF, 투자CF, 재무CF, 현금증감 이렇게 7개만 가져옴

#   1. 시가총액 순위별로 회사 Code가 입력되어 있는 파일을 참조할 준비를 한다
#   2. 그 파일에 불러온 값을 저장할 준비도 한다
from openpyxl import load_workbook
path = "c:/Users/msmk1/OneDrive/Routine/Auto/BPC_CF.xlsx"
wb = load_workbook(path)
ws = wb["Rank"]
db = wb["Data"]


#   3. 필요한 라이브러리들 불러오고
import requests
from bs4 import BeautifulSoup

#   한 회사당 가져올 항목이 몇개인지 들어가 있는 변수
#   요 변수만큼 칸을 띄어서 엑셀에 입력하게 된다.
no_eachline = 7


#   4. 회사 Code와 이름을 불러온다
for x in range(1,ws.max_row+1):
    company_Code = ws.cell(row=x, column=3).value
    company_Name = ws.cell(row=x, column=2).value
       

#   5. FnGuide 해당 페이지 접속해서 데이터 가져온다
    url = "https://comp.fnguide.com/SVO2/ASP/SVD_Finance.asp?pGB=1&gicode=A" + company_Code + "&cID=&MenuYn=Y&ReportGB=&NewMenuID=103&stkGb=701"
    res = requests.get(url)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")


#   6-1. PL
#        divSonikY 를 id로 하는 테이블이 PL 나온 테이블임
    data_rows = soup.find("div", attrs={"id":"divSonikY"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    i = 0
    for row in data_rows:
        i = i + 1
        C_head = row.find("th").get_text()
        if C_head == """\n매출액\n""":
            int매출액 = i
            db.cell(row=1 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=1 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=1 + no_eachline * (x-1), column=3, value= "매출")
        elif C_head =="""\n순이자손익\n""":
            int매출액 = i
            db.cell(row=1 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=1 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=1 + no_eachline * (x-1), column=3, value= "매출")
        elif C_head =="""\n영업수익\n""":
            int매출액 = i
            db.cell(row=1 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=1 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=1 + no_eachline * (x-1), column=3, value= "매출")
        elif C_head == """\n영업이익\n""":
            int영업이익 = i
            db.cell(row=2 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=2 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=2 + no_eachline * (x-1), column=3, value= "영업이익")
        elif C_head == """\n당기순이익\n""":
            int당기순이익 = i
            db.cell(row=3 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=3 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=3 + no_eachline * (x-1), column=3, value= "순이익")
        else:
            dumb_num = 0


#   6-2. 연결/연간
    i = 0
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        if i == int매출액:
            db.cell(row=1 + no_eachline * (x-1), column=5, value = data[0])
            db.cell(row=1 + no_eachline * (x-1), column=6, value = data[1])
            db.cell(row=1 + no_eachline * (x-1), column=7, value = data[2])
            db.cell(row=1 + no_eachline * (x-1), column=8, value = data[3])
        elif i == int영업이익:
            db.cell(row=2 + no_eachline * (x-1), column=5, value = data[0])
            db.cell(row=2 + no_eachline * (x-1), column=6, value = data[1])
            db.cell(row=2 + no_eachline * (x-1), column=7, value = data[2])
            db.cell(row=2 + no_eachline * (x-1), column=8, value = data[3])
        elif i == int당기순이익:
            db.cell(row=3 + no_eachline * (x-1), column=5, value = data[0])
            db.cell(row=3 + no_eachline * (x-1), column=6, value = data[1])
            db.cell(row=3 + no_eachline * (x-1), column=7, value = data[2])
            db.cell(row=3 + no_eachline * (x-1), column=8, value = data[3])
        else:
            dumb_num = 0


#   6-3. 연결/분기
    data_rows = soup.find("div", attrs={"id":"divSonikQ"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    i = 0
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        if i == int매출액:
            db.cell(row=1 + no_eachline * (x-1), column=10, value = data[0])
            db.cell(row=1 + no_eachline * (x-1), column=11, value = data[1])
            db.cell(row=1 + no_eachline * (x-1), column=12, value = data[2])
            db.cell(row=1 + no_eachline * (x-1), column=13, value = data[3])
        elif i == int영업이익:
            db.cell(row=2 + no_eachline * (x-1), column=10, value = data[0])
            db.cell(row=2 + no_eachline * (x-1), column=11, value = data[1])
            db.cell(row=2 + no_eachline * (x-1), column=12, value = data[2])
            db.cell(row=2 + no_eachline * (x-1), column=13, value = data[3])
        elif i == int당기순이익:
            db.cell(row=3 + no_eachline * (x-1), column=10, value = data[0])
            db.cell(row=3 + no_eachline * (x-1), column=11, value = data[1])
            db.cell(row=3 + no_eachline * (x-1), column=12, value = data[2])
            db.cell(row=3 + no_eachline * (x-1), column=13, value = data[3])
        else:
            dumb_num = 0


#   8-1. CF
    k = i
    data_rows = soup.find("div", attrs={"id":"divCashY"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    for row in data_rows:
        i = i + 1
        C_head = row.find("th").get_text()
        if C_head == """\n영업활동으로인한현금흐름\n""":
            int영업CF = i
            db.cell(row=4 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=4 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=4 + no_eachline * (x-1), column=3, value= "영업CF")
        elif C_head == """\n투자활동으로인한현금흐름\n""":
            int투자CF = i
            db.cell(row=5 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=5 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=5 + no_eachline * (x-1), column=3, value= "투자CF")
        elif C_head == """\n재무활동으로인한현금흐름\n""":
            int재무CF = i
            db.cell(row=6 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=6 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=6 + no_eachline * (x-1), column=3, value= "재무CF")
        elif C_head == """\n현금및현금성자산의증가\n""":
            int현금증감 = i
            db.cell(row=7 + no_eachline * (x-1), column=1, value= company_Code)
            db.cell(row=7 + no_eachline * (x-1), column=2, value= company_Name)
            db.cell(row=7 + no_eachline * (x-1), column=3, value= "현금증감")
        else:
            dumb_num = 0


#   8-2. 연결/연간
    i = k
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        if i == int영업CF:
            db.cell(row=4 + no_eachline * (x-1), column=5, value= data[0])
            db.cell(row=4 + no_eachline * (x-1), column=6, value= data[1])
            db.cell(row=4 + no_eachline * (x-1), column=7, value= data[2])
            db.cell(row=4 + no_eachline * (x-1), column=8, value= data[3])
        elif i == int투자CF:
            db.cell(row=5 + no_eachline * (x-1), column=5, value= data[0])
            db.cell(row=5 + no_eachline * (x-1), column=6, value= data[1])
            db.cell(row=5 + no_eachline * (x-1), column=7, value= data[2])
            db.cell(row=5 + no_eachline * (x-1), column=8, value= data[3])
        elif i == int재무CF:
            db.cell(row=6 + no_eachline * (x-1), column=5, value= data[0])
            db.cell(row=6 + no_eachline * (x-1), column=6, value= data[1])
            db.cell(row=6 + no_eachline * (x-1), column=7, value= data[2])
            db.cell(row=6 + no_eachline * (x-1), column=8, value= data[3])
        elif i == int현금증감:
            db.cell(row=7 + no_eachline * (x-1), column=5, value= data[0])
            db.cell(row=7 + no_eachline * (x-1), column=6, value= data[1])
            db.cell(row=7 + no_eachline * (x-1), column=7, value= data[2])
            db.cell(row=7 + no_eachline * (x-1), column=8, value= data[3])
        else:
            dumb_num = 0


#   8-3. 연결/분기
    i = k
    data_rows = soup.find("div", attrs={"id":"divCashQ"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        if i == int영업CF:
            db.cell(row=4 + no_eachline * (x-1), column=10, value= data[0])
            db.cell(row=4 + no_eachline * (x-1), column=11, value= data[1])
            db.cell(row=4 + no_eachline * (x-1), column=12, value= data[2])
            db.cell(row=4 + no_eachline * (x-1), column=13, value= data[3])
        elif i == int투자CF:
            db.cell(row=5 + no_eachline * (x-1), column=10, value= data[0])
            db.cell(row=5 + no_eachline * (x-1), column=11, value= data[1])
            db.cell(row=5 + no_eachline * (x-1), column=12, value= data[2])
            db.cell(row=5 + no_eachline * (x-1), column=13, value= data[3])
        elif i == int재무CF:
            db.cell(row=6 + no_eachline * (x-1), column=10, value= data[0])
            db.cell(row=6 + no_eachline * (x-1), column=11, value= data[1])
            db.cell(row=6 + no_eachline * (x-1), column=12, value= data[2])
            db.cell(row=6 + no_eachline * (x-1), column=13, value= data[3])
        elif i == int현금증감:
            db.cell(row=7 + no_eachline * (x-1), column=10, value= data[0])
            db.cell(row=7 + no_eachline * (x-1), column=11, value= data[1])
            db.cell(row=7 + no_eachline * (x-1), column=12, value= data[2])
            db.cell(row=7 + no_eachline * (x-1), column=13, value= data[3])
        else:
            dumb_num = 0

#   진행된 내용 표시하고, 200개 검토될 때마다 저장한다.
    print("진행률 : " + str(x) + "/" + str(ws.max_row+1) + ", " + company_Name)
    if x % 200 == 0:
        wb.save(path)

wb.save(path)