# '21.09.12 일요일에 최초로 완성하고, 1000개 종목의 데이터를 가져와 저장해 봤음

#   1. 시가총액 순위별로 회사 Code가 입력되어 있는 파일을 참조할 준비를 한다
#   2. 그 파일에 불러온 값을 저장할 준비도 한다
from openpyxl import load_workbook
path = "c:/Users/msmk1/OneDrive/Routine/Auto/BPC.xlsx"
wb = load_workbook(path)
ws = wb["Rank"]
db = wb["Data"]


#   3. 필요한 라이브러리들 불러오고
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm


# 슬랙 메시지 전달
def post_message(token, channel, text):
    response = requests.post("https://slack.com/api/chat.postMessage",
        headers={"Authorization": "Bearer "+token},
        data={"channel": channel,"text": text}
    )
    print(response)

myToken = "xoxb-2483926020578-2469328976551-Qcu5xUqVzxqSsPeQ0Mv5RxHz"


#try:
no_eachline = 303
#   4. 회사 Code와 이름을 불러온다
#for x in tqdm(range(1,ws.max_row+1)):
for x in range(1,ws.max_row+1):
    company_Code = ws.cell(row=x, column=3).value
    company_Name = ws.cell(row=x, column=2).value
    progress = x, company_Name
    
    

#   5. FnGuide 해당 페이지 접속해서 데이터 가져온다
    url = "https://comp.fnguide.com/SVO2/ASP/SVD_Finance.asp?pGB=1&gicode=A" + company_Code + "&cID=&MenuYn=Y&ReportGB=&NewMenuID=103&stkGb=701"
    res = requests.get(url)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")


#   6-1. PL
    data_rows = soup.find("div", attrs={"id":"divSonikY"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    i = 0
    for row in data_rows:
        i = i + 1
        C_head = row.find("th").get_text()
        db.cell(row=i + no_eachline * (x-1), column=1, value= company_Code)
        db.cell(row=i + no_eachline * (x-1), column=2, value= company_Name)
        db.cell(row=i + no_eachline * (x-1), column=3, value= "PL")
        db.cell(row=i + no_eachline * (x-1), column=5, value= C_head)
    #wb.save(path)
#   6-2. 연결/연간
    i = 0
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        db.cell(row=i + no_eachline * (x-1), column=6, value = data[0])
        db.cell(row=i + no_eachline * (x-1), column=7, value = data[1])
        db.cell(row=i + no_eachline * (x-1), column=8, value = data[2])
        db.cell(row=i + no_eachline * (x-1), column=9, value = data[3])
        db.cell(row=i + no_eachline * (x-1), column=10, value = data[4])
        db.cell(row=i + no_eachline * (x-1), column=11, value = data[5])
    wb.save(path)        
#   6-3. 연결/분기
    data_rows = soup.find("div", attrs={"id":"divSonikQ"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    i = 0
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        db.cell(row=i + no_eachline * (x-1), column=13, value = data[0])
        db.cell(row=i + no_eachline * (x-1), column=14, value = data[1])
        db.cell(row=i + no_eachline * (x-1), column=15, value = data[2])
        db.cell(row=i + no_eachline * (x-1), column=16, value = data[3])
        db.cell(row=i + no_eachline * (x-1), column=17, value = data[4])
        db.cell(row=i + no_eachline * (x-1), column=18, value = data[5])
    wb.save(path)


#   7-1. BS
    j = i
    data_rows = soup.find("div", attrs={"id":"divDaechaY"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    for row in data_rows:
        i = i + 1
        C_head = row.find("th").get_text()
        db.cell(row=i + no_eachline * (x-1), column=1, value= company_Code)
        db.cell(row=i + no_eachline * (x-1), column=2, value= company_Name)
        db.cell(row=i + no_eachline * (x-1), column=3, value= "BS")
        db.cell(row=i + no_eachline * (x-1), column=5, value= C_head)
    #wb.save(path)
#   7-2. 연결/연간
    i = j
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        db.cell(row=i + no_eachline * (x-1), column=6, value = data[0])
        db.cell(row=i + no_eachline * (x-1), column=7, value = data[1])
        db.cell(row=i + no_eachline * (x-1), column=8, value = data[2])
        db.cell(row=i + no_eachline * (x-1), column=9, value = data[3])
    wb.save(path)        
#   7-3. 연결/분기
    i = j
    data_rows = soup.find("div", attrs={"id":"divDaechaQ"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        db.cell(row=i + no_eachline * (x-1), column=13, value = data[0])
        db.cell(row=i + no_eachline * (x-1), column=14, value = data[1])
        db.cell(row=i + no_eachline * (x-1), column=15, value = data[2])
        db.cell(row=i + no_eachline * (x-1), column=16, value = data[3])
    wb.save(path)


#   8-1. CF
    k = i
    data_rows = soup.find("div", attrs={"id":"divCashY"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    for row in data_rows:
        i = i + 1
        C_head = row.find("th").get_text()
        db.cell(row=i + no_eachline * (x-1), column=1, value= company_Code)
        db.cell(row=i + no_eachline * (x-1), column=2, value= company_Name)
        db.cell(row=i + no_eachline * (x-1), column=3, value= "CF")
        db.cell(row=i + no_eachline * (x-1), column=5, value= C_head)
    #wb.save(path)
#   8-2. 연결/연간
    i = k
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        db.cell(row=i + no_eachline * (x-1), column=6, value = data[0])
        db.cell(row=i + no_eachline * (x-1), column=7, value = data[1])
        db.cell(row=i + no_eachline * (x-1), column=8, value = data[2])
        db.cell(row=i + no_eachline * (x-1), column=9, value = data[3])
    wb.save(path)        
#   8-3. 연결/분기
    i = k
    data_rows = soup.find("div", attrs={"id":"divCashQ"}).find("table", attrs={"class":"us_table_ty1 h_fix zigbg_no"}).find("tbody").find_all("tr")
    for row in data_rows:
        i = i + 1
        columns = row.find_all("td")
        data = [column.get_text() for column in columns]
        db.cell(row=i + no_eachline * (x-1), column=13, value = data[0])
        db.cell(row=i + no_eachline * (x-1), column=14, value = data[1])
        db.cell(row=i + no_eachline * (x-1), column=15, value = data[2])
        db.cell(row=i + no_eachline * (x-1), column=16, value = data[3])
    wb.save(path)
    
    print(progress)
    post_message(myToken,"#stock_test",progress)

post_message(myToken,"#stock_test","Finished")

#except:
#    post_message(myToken,"#stock_test","Error")
