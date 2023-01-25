from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()                  # 새로운 sheet 생성, "기본 이름으로"
ws.title = "MySheet"                    # Sheet 이름 변경
ws.sheet_properties.tabColor = "33cc33"

ws1 = wb.create_sheet("YourSheet")      # 새로운 Sheet 생성, 주어진 이름으로
ws2 = wb.create_sheet("NewSheet", 2)    # 새로운 Sheet 생성, 주어진 이름으로, 2번째 인덱스 위치에

new_ws = wb["NewSheet"]                 # 딕셔너리 형태로 Sheet에 접근

print(wb.sheetnames)                    # 모든 시트 이름 확인


# Sheet 내용 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)      # workbook 안에 있는 new_ws 워크시트 복사
target.title = "Copied Sheet"

print(wb.sheetnames)

wb.save("sample.xlsx")