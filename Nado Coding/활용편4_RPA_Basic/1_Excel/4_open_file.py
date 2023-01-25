from openpyxl import load_workbook      # 파일 불러오기 위해 필요한 라이브러리
wb = load_workbook("sample.xlsx")       # sample 파일에서 워크북 불러옴
ws = wb.active                          # 현재 workbook의 활성화된 Sheet

# cell 데이터 불러오기
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x, column=y).value, end=" ")  # 뒤에 end=" " 해주지 않으면 값하나마다 줄바꿈 된다
    print()                                             # print() 하면 줄바꿈 된다


print("-------------------------------")                                           


# cell 이 몇개인지 모를때
for x in range(1,ws.max_row+1):
    for y in range(1,ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()                                           


