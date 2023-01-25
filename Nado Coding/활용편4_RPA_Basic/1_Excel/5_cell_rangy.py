from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 데이터 한줄씩 넣기
ws.append(["번호","영어","수학"])                        # A, B, C

# 10줄의 데이터 넣기
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

print("------------1--------------")

col_B = ws["B"]                                         # 영어 column만 가져오기
print(col_B)

for cell in col_B:
    print(cell.value)

print("------------2--------------")

col_range = ws["B:C"]                                   # 영어, 수학 column 함께 가져오기
for cols in col_range:
    for cell in cols:
            print(cell.value)

print("------------3--------------")

row_title = ws[1]                                       # 첫번째 row 가져오기
for cell in row_title:
    print(cell.value)

print("------------4--------------")

row_range = ws[2:6]                                     # 2번째 줄에서 6번째 줄까지 가져오기
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

print("------------5--------------")

row_range = ws[2:ws.max_row]                            # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

print("------------6--------------")

from openpyxl.utils.cell import coordinate_from_string  # cell의 좌표정보 가져오기

row_range = ws[1:ws.max_row]                            
for rows in row_range:
    for cell in rows:
        print(cell.coordinate, end=" ")
    print()

print("------------7--------------")

row_range = ws[1:ws.max_row]                            
for rows in row_range:
    for cell in rows:
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")
    print()

print("------------8--------------")

row_range = ws[1:ws.max_row]                            
for rows in row_range:
    for cell in rows:
        xy = coordinate_from_string(cell.coordinate)
        print(xy[0], end="")
        print(xy[1], end=" ")
    print()

print("------------9--------------")
print(tuple(ws.rows))                   # 한 줄씩 튜플로 묶어서 가져옴


print("------------10--------------")
print(tuple(ws.columns))                # 한 칼럼씩 튜플로 묶어서 가져옴


print("------------11--------------")
for row in tuple(ws.rows):
    print(row)

for row in tuple(ws.rows):
    print(row[2].value)

print("------------12--------------")
for column in tuple(ws.columns):
    print(column)

for column in tuple(ws.columns):
    print(column[0].value)


print("------------13--------------")
for row in ws.iter_rows():  # 전체 row 에 대해서 반복하면서..
    print(row[1].value)

for column in ws.iter_cols(): 
    print(column[0].value)


print("------------14--------------")
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):  # 전체 row 에 대해서 반복하면서..
    print(row[0].value, row[1].value)
    print(row)

print("------------15--------------")
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):  # 전체 row 에 대해서 반복하면서..
    print(col)


wb.save("sample.xlsx")


#1:02:25